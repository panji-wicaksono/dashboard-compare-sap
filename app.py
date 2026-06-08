import os, io, json, traceback, sqlite3, subprocess, threading
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024

DB_PATH = os.path.join(os.path.dirname(__file__), "dashboard.db")

# ── Database ───────────────────────────────────────────────────────────────────

SCHEMA = [
    """CREATE TABLE IF NOT EXISTS aufm (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        MANDT TEXT, MBLNR TEXT, MJAHR TEXT, ZEILE TEXT,
        AUFNR TEXT, MATNR TEXT, BWART TEXT, MENGE REAL,
        MEINS TEXT, BLDAT TEXT, WERKS TEXT, LGORT TEXT
    )""",
    """CREATE TABLE IF NOT EXISTS caufv (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        AUFNR TEXT UNIQUE, AUART TEXT, GSTRP TEXT, WERKS TEXT
    )""",
    """CREATE TABLE IF NOT EXISTS makt (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        MATNR TEXT UNIQUE, MAKTX TEXT
    )""",
    """CREATE TABLE IF NOT EXISTS prodsys (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        SAPID TEXT, AUFNR TEXT, MATNR TEXT, MENGE REAL, CONVM REAL, ZDATA1 REAL, MEINS TEXT,
        GSTRP TEXT, AFRUD_ISDD TEXT, WERKS TEXT, AUART TEXT, KEY_STATUS TEXT
    )""",
    """CREATE TABLE IF NOT EXISTS upload_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tbl TEXT, filename TEXT, rows INTEGER,
        ts TEXT DEFAULT (datetime('now','localtime'))
    )""",
    "CREATE INDEX IF NOT EXISTS idx_aufm_date   ON aufm(BLDAT)",
    "CREATE INDEX IF NOT EXISTS idx_aufm_bwart  ON aufm(BWART)",
    "CREATE INDEX IF NOT EXISTS idx_aufm_order  ON aufm(AUFNR)",
    "CREATE INDEX IF NOT EXISTS idx_caufv_order ON caufv(AUFNR)",
    "CREATE INDEX IF NOT EXISTS idx_makt_mat    ON makt(MATNR)",
]


def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        for stmt in SCHEMA:
            conn.execute(stmt)

        # Column migrations
        for col_sql in [
            "ALTER TABLE prodsys ADD COLUMN GSTRP TEXT",
            "ALTER TABLE prodsys ADD COLUMN SAPID TEXT",
            "ALTER TABLE prodsys ADD COLUMN CONVM REAL",
            "ALTER TABLE prodsys ADD COLUMN ZDATA1 REAL",
            "ALTER TABLE aufm ADD COLUMN MANDT TEXT",
            "ALTER TABLE aufm ADD COLUMN MJAHR TEXT",
            "ALTER TABLE aufm ADD COLUMN ZEILE TEXT",
            "ALTER TABLE aufm ADD COLUMN LGORT TEXT",
        ]:
            try:
                conn.execute(col_sql)
            except Exception:
                pass

        # Drop old AUFM unique index if it was on the wrong key, recreate on SAP PK
        try:
            conn.execute("DROP INDEX IF EXISTS idx_aufm_unique")
        except Exception:
            pass

        # Unique indexes — deduplicate existing rows first, then create index
        _dedup_and_index(conn, "aufm",
            "MANDT, MBLNR, MJAHR, ZEILE",
            "idx_aufm_unique")
        # Drop old prodsys unique index if it was on the wrong key
        try:
            conn.execute("DROP INDEX IF EXISTS idx_prodsys_unique")
        except Exception:
            pass
        _dedup_and_index(conn, "prodsys",
            "SAPID, WERKS",
            "idx_prodsys_unique")

        conn.execute("CREATE INDEX IF NOT EXISTS idx_prd_gstrp ON prodsys(GSTRP)")
        conn.commit()


def _dedup_and_index(conn, tbl, key_cols, idx_name):
    try:
        conn.execute(
            f"DELETE FROM {tbl} WHERE rowid NOT IN "
            f"(SELECT MIN(rowid) FROM {tbl} GROUP BY {key_cols})"
        )
        conn.execute(
            f"CREATE UNIQUE INDEX IF NOT EXISTS {idx_name} ON {tbl}({key_cols})"
        )
    except Exception:
        pass


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


class _RawFile:
    """Wrapper agar file lokal bisa dipakai seperti Flask FileStorage oleh read_sap_file."""
    def __init__(self, path):
        self.filename = os.path.basename(path)
        self._path = path
    def read(self):
        with open(self._path, "rb") as fh:
            return fh.read()


_auto = {"running": False, "log": [], "error": None, "done": False}


# ── Parsers ────────────────────────────────────────────────────────────────────

def read_sap_file(file_storage):
    """Read SAP export (Excel/CSV). Skips the description header row SAP adds."""
    name = file_storage.filename.lower()
    raw = file_storage.read()

    if name.endswith((".xlsx", ".xls")):
        try:
            df = pd.read_excel(io.BytesIO(raw), dtype=str)
        except Exception:
            # SAP sering export tab-separated text dengan ekstensi .xls
            text = raw.decode("utf-8-sig", errors="replace")
            df = pd.read_csv(io.StringIO(text), sep="\t", dtype=str)
    elif name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        first = text.split("\n")[0]
        sep = "\t" if "\t" in first else (";" if first.count(";") > first.count(",") else ",")
        df = pd.read_csv(io.StringIO(text), sep=sep, dtype=str)
    else:
        raise ValueError("Format tidak didukung. Gunakan .xlsx, .xls, atau .csv")

    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")

    # Skip the SAP description row (row 0 has 'Client' in MANDT instead of a number)
    if len(df) > 0:
        first_mandt = str(df.iloc[0].get("MANDT", "")).strip()
        if not first_mandt.isdigit():
            df = df.iloc[1:].reset_index(drop=True)

    df = df.dropna(how="all").reset_index(drop=True)
    return df


def parse_num(val):
    if pd.isna(val) or str(val).strip() in ("", "-", "nan", "None"):
        return 0.0
    s = str(val).strip()
    if "," in s and "." in s:
        # European: 1.234,56 or US: 1,234.56
        s = s.replace(".", "").replace(",", ".") if s.rindex(",") > s.rindex(".") else s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        s = s.replace(",", ".") if len(parts) == 2 and len(parts[1]) <= 3 and parts[1].isdigit() else s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date_dmy(val):
    """DD.MM.YYYY → YYYY-MM-DD"""
    s = str(val).strip() if not pd.isna(val) else ""
    if not s:
        return None
    try:
        return pd.to_datetime(s, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return None


def parse_date_iso(val):
    """ISO datetime string → YYYY-MM-DD"""
    s = str(val).strip() if not pd.isna(val) else ""
    if not s:
        return None
    try:
        return pd.to_datetime(s).strftime("%Y-%m-%d")
    except Exception:
        return None


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ── Insert helpers (dipakai oleh upload routes DAN automation) ─────────────────

def _insert_aufm(f):
    df = read_sap_file(f)
    required = ["MANDT", "MBLNR", "MJAHR", "ZEILE", "AUFNR", "MATNR", "BWART", "MENGE", "MEINS", "BLDAT", "WERKS"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Kolom tidak ditemukan: {missing}. Kolom tersedia: {list(df.columns)}")

    df["MENGE"] = df["MENGE"].apply(parse_num)
    df["BLDAT"] = df["BLDAT"].apply(parse_date_dmy)
    df = df[df["BWART"].isin(["531", "532", "888", "889"])]

    rows = [
        (
            str(r.get("MANDT", "") or "").strip(),
            str(r.get("MBLNR", "") or "").strip(),
            str(r.get("MJAHR", "") or "").strip(),
            str(r.get("ZEILE", "") or "").strip(),
            str(r.get("AUFNR", "") or "").strip(),
            str(r.get("MATNR", "") or "").strip(),
            str(r.get("BWART", "") or "").strip(),
            r.get("MENGE", 0),
            str(r.get("MEINS", "") or "").strip(),
            r.get("BLDAT"),
            str(r.get("WERKS", "") or "").strip(),
            str(r.get("LGORT", "") or "").strip() if "LGORT" in df.columns else "",
        )
        for _, r in df.iterrows()
        if str(r.get("AUFNR", "") or "").strip()
    ]

    with get_db() as conn:
        before = conn.execute("SELECT COUNT(*) FROM aufm").fetchone()[0]
        conn.executemany(
            "INSERT OR IGNORE INTO aufm (MANDT,MBLNR,MJAHR,ZEILE,AUFNR,MATNR,BWART,MENGE,MEINS,BLDAT,WERKS,LGORT) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        inserted = conn.execute("SELECT COUNT(*) FROM aufm").fetchone()[0] - before
        conn.execute("INSERT INTO upload_log (tbl,filename,rows) VALUES (?,?,?)", ("aufm", f.filename, inserted))
        conn.commit()

    return inserted, len(rows) - inserted


def _insert_caufv(f):
    df = read_sap_file(f)
    missing = [c for c in ["AUFNR", "AUART"] if c not in df.columns]
    if missing:
        raise ValueError(f"Kolom tidak ditemukan: {missing}. Tersedia: {list(df.columns)}")

    gstrp_col = next((c for c in ["GSTRP", "FTRMS", "GSTRS"] if c in df.columns), None)

    rows = [
        (
            str(r.get("AUFNR", "") or "").strip(),
            str(r.get("AUART", "") or "").strip(),
            parse_date_dmy(r.get(gstrp_col)) if gstrp_col else None,
            str(r.get("WERKS", "") or "").strip(),
        )
        for _, r in df.iterrows()
        if str(r.get("AUFNR", "") or "").strip() and str(r.get("AUART", "") or "").strip()
    ]

    with get_db() as conn:
        before = conn.execute("SELECT COUNT(*) FROM caufv").fetchone()[0]
        conn.executemany("INSERT OR IGNORE INTO caufv (AUFNR,AUART,GSTRP,WERKS) VALUES (?,?,?,?)", rows)
        inserted = conn.execute("SELECT COUNT(*) FROM caufv").fetchone()[0] - before
        conn.execute("INSERT INTO upload_log (tbl,filename,rows) VALUES (?,?,?)", ("caufv", f.filename, inserted))
        conn.commit()

    return inserted, len(rows) - inserted


def _insert_makt(f):
    df = read_sap_file(f)
    missing = [c for c in ["MATNR", "MAKTX"] if c not in df.columns]
    if missing:
        raise ValueError(f"Kolom tidak ditemukan: {missing}. Tersedia: {list(df.columns)}")

    rows = [
        (str(r.get("MATNR", "") or "").strip(), str(r.get("MAKTX", "") or "").strip())
        for _, r in df.iterrows()
        if str(r.get("MATNR", "") or "").strip()
    ]

    with get_db() as conn:
        before = conn.execute("SELECT COUNT(*) FROM makt").fetchone()[0]
        conn.executemany("INSERT OR IGNORE INTO makt (MATNR,MAKTX) VALUES (?,?)", rows)
        inserted = conn.execute("SELECT COUNT(*) FROM makt").fetchone()[0] - before
        conn.execute("INSERT INTO upload_log (tbl,filename,rows) VALUES (?,?,?)", ("makt", f.filename, inserted))
        conn.commit()

    return inserted, len(rows) - inserted


def _insert_prodsys(f):
    df = read_sap_file(f)
    required = ["SAPID", "CAUFV_AUFN", "CAUFV_WERK", "MSEG_MATNR", "MSEG_MENGE", "MSEG_CONVM", "ZDATA1"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Kolom tidak ditemukan: {missing}. Tersedia: {list(df.columns)}")

    df["MSEG_MENGE"] = df["MSEG_MENGE"].apply(parse_num)
    df["MSEG_CONVM"] = df["MSEG_CONVM"].apply(parse_num)
    df["ZDATA1"]     = df["ZDATA1"].apply(parse_num)

    if "CAUFV_GSTRP" in df.columns:
        df["_GSTRP"] = df["CAUFV_GSTRP"].apply(parse_date_dmy)
    elif "AFRUD_ISDD" in df.columns:
        df["_GSTRP"] = df["AFRUD_ISDD"].apply(parse_date_iso)
    else:
        df["_GSTRP"] = None

    df["_AFRUD"] = df["AFRUD_ISDD"].apply(parse_date_iso) if "AFRUD_ISDD" in df.columns else None

    rows = [
        (
            str(r.get("SAPID", "") or "").strip(),
            str(r.get("CAUFV_AUFN", "") or "").strip(),
            str(r.get("MSEG_MATNR", "") or "").strip(),
            r.get("MSEG_MENGE", 0),
            r.get("MSEG_CONVM", 0),
            r.get("ZDATA1", 0),
            str(r.get("MSEG_MEINS", "") or "").strip(),
            r.get("_GSTRP"),
            r.get("_AFRUD"),
            str(r.get("CAUFV_WERK", "") or "").strip(),
            str(r.get("CAUFV_AUAR", "") or "").strip(),
            str(r.get("KEY_STATUS", "") or "").strip(),
        )
        for _, r in df.iterrows()
        if str(r.get("SAPID", "") or "").strip()
    ]

    src = "CAUFV_GSTRP" if "CAUFV_GSTRP" in df.columns else "AFRUD_ISDD (fallback)"

    with get_db() as conn:
        before = conn.execute("SELECT COUNT(*) FROM prodsys").fetchone()[0]
        conn.executemany(
            "INSERT OR IGNORE INTO prodsys (SAPID,AUFNR,MATNR,MENGE,CONVM,ZDATA1,MEINS,GSTRP,AFRUD_ISDD,WERKS,AUART,KEY_STATUS) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        inserted = conn.execute("SELECT COUNT(*) FROM prodsys").fetchone()[0] - before
        conn.execute("INSERT INTO upload_log (tbl,filename,rows) VALUES (?,?,?)", ("prodsys", f.filename, inserted))
        conn.commit()

    return inserted, len(rows) - inserted, src


# ── Upload APIs ────────────────────────────────────────────────────────────────

@app.route("/api/upload/aufm", methods=["POST"])
def upload_aufm():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "File diperlukan"}), 400
    try:
        inserted, skipped = _insert_aufm(f)
        return jsonify({"ok": True, "rows": inserted, "skipped": skipped, "filename": f.filename})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/upload/caufv", methods=["POST"])
def upload_caufv():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "File diperlukan"}), 400
    try:
        inserted, skipped = _insert_caufv(f)
        return jsonify({"ok": True, "rows": inserted, "skipped": skipped, "filename": f.filename})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/upload/makt", methods=["POST"])
def upload_makt():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "File diperlukan"}), 400
    try:
        inserted, skipped = _insert_makt(f)
        return jsonify({"ok": True, "rows": inserted, "skipped": skipped, "filename": f.filename})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/upload/prodsys", methods=["POST"])
def upload_prodsys():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "File diperlukan"}), 400
    try:
        inserted, skipped, src = _insert_prodsys(f)
        return jsonify({"ok": True, "rows": inserted, "skipped": skipped, "filename": f.filename, "date_source": src})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── Automation ─────────────────────────────────────────────────────────────────

def _run_automation(date_dmy, date_iso):
    macro_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "macro sap")
    data_dir  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data upload")

    def log(msg):
        _auto["log"].append(msg)

    try:
        # 1. Download CAUFV (filter GSTRP = tanggal dipilih, format DD.MM.YYYY)
        subprocess.run(["wscript", os.path.join(macro_dir, "caufv.vbs"), date_dmy], check=True)
        log(f"CAUFV didownload ({date_dmy})")

        # 2. Baca AUFNR dari caufv.xls → salin ke clipboard
        df_c = read_sap_file(_RawFile(os.path.join(data_dir, "caufv.xls")))
        aufnr_list = df_c["AUFNR"].dropna().str.strip().unique().tolist()
        aufnr_str  = "\n".join(aufnr_list)
        subprocess.run("clip", input=aufnr_str.encode("utf-8"), shell=True, check=True)
        log(f"{len(aufnr_list)} AUFNR disalin ke clipboard")

        # 3. Download AUFM (menggunakan clipboard AUFNR)
        subprocess.run(["wscript", os.path.join(macro_dir, "aufm.vbs")], check=True)
        log("AUFM didownload dari SAP")

        # 4. Download ZPPCPFINT_GRAUTO (filter date = YYYY-MM-DD*)
        subprocess.run(["wscript", os.path.join(macro_dir, "zppcpfint_grauto.vbs"), date_iso + "*"], check=True)
        log(f"ZPPCPFINT_GRAUTO didownload ({date_iso})")

        # 5. Upload CAUFV ke database
        ins, skp = _insert_caufv(_RawFile(os.path.join(data_dir, "caufv.xls")))
        log(f"CAUFV tersimpan: {ins} baris baru, {skp} duplikat")

        # 6. Upload AUFM ke database
        ins, skp = _insert_aufm(_RawFile(os.path.join(data_dir, "aufm.xls")))
        log(f"AUFM tersimpan: {ins} baris baru, {skp} duplikat")

        # 7. Upload Prodsys ke database
        ins, skp, _ = _insert_prodsys(_RawFile(os.path.join(data_dir, "zppcpfint_grauto.xls")))
        log(f"Prodsys tersimpan: {ins} baris baru, {skp} duplikat")

        _auto["done"] = True

    except Exception as e:
        _auto["error"] = str(e)
    finally:
        _auto["running"] = False


@app.route("/api/automation/run", methods=["POST"])
def automation_run():
    if _auto["running"]:
        return jsonify({"error": "Automasi sedang berjalan"}), 409

    data     = request.get_json(silent=True) or {}
    date_str = data.get("date", "").strip()

    if date_str:
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Format tanggal tidak valid. Gunakan YYYY-MM-DD"}), 400
    else:
        dt = datetime.now()

    date_dmy = dt.strftime("%d.%m.%Y")
    date_iso = dt.strftime("%Y-%m-%d")

    _auto.update({"running": True, "log": [], "error": None, "done": False})
    threading.Thread(target=_run_automation, args=(date_dmy, date_iso), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/automation/status")
def automation_status():
    return jsonify({
        "running": _auto["running"],
        "log":     _auto["log"],
        "error":   _auto["error"],
        "done":    _auto["done"],
    })


@app.route("/api/upload/status")
def upload_status():
    try:
        with get_db() as conn:
            result = {}
            for tbl in ["aufm", "caufv", "makt", "prodsys"]:
                cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
                log = conn.execute(
                    "SELECT filename, rows, ts FROM upload_log WHERE tbl=? ORDER BY id DESC LIMIT 1", (tbl,)
                ).fetchone()
                result[tbl] = {
                    "rows": cnt,
                    "file": log["filename"] if log else None,
                    "log_rows": log["rows"] if log else None,
                    "ts": log["ts"] if log else None,
                }
            return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Compare APIs ───────────────────────────────────────────────────────────────

@app.route("/api/dates")
def get_dates():
    """Available GSTRP dates in prodsys."""
    try:
        with get_db() as conn:
            rows = conn.execute(
                "SELECT DISTINCT GSTRP FROM prodsys WHERE GSTRP IS NOT NULL ORDER BY GSTRP DESC"
            ).fetchall()
            return jsonify({"dates": [r["GSTRP"] for r in rows]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/compare")
def compare():
    date = request.args.get("date", "").strip()
    if not date:
        return jsonify({"error": "Parameter date diperlukan"}), 400

    try:
        with get_db() as conn:
            # SAP: AUFM untuk order yang ada di Prodsys pada GSTRP tersebut.
            # Filter by AUFNR (bukan BLDAT) karena tanggal posting SAP bisa berbeda dgn GSTRP.
            # 531/888 = GR (positive), 532/889 = cancellation (dikurangi)
            sap_df = pd.read_sql_query(
                """
                SELECT m.AUFNR, m.MATNR,
                       SUM(CASE WHEN m.BWART IN ('531','888') THEN m.MENGE ELSE -m.MENGE END) AS qty_sap,
                       MAX(m.MEINS)              AS meins,
                       MAX(m.WERKS)              AS werks,
                       MAX(COALESCE(c.AUART,'')) AS auart
                FROM aufm m
                LEFT JOIN caufv c ON m.AUFNR = c.AUFNR
                WHERE m.BWART IN ('531','532','888','889')
                  AND (m.LGORT IS NULL OR m.LGORT != 'SHUV')
                  AND m.AUFNR IN (
                      SELECT DISTINCT AUFNR FROM prodsys WHERE GSTRP = :dt
                      UNION
                      SELECT DISTINCT AUFNR FROM caufv WHERE GSTRP = :dt
                  )
                GROUP BY m.AUFNR, m.MATNR
                """,
                conn, params={"dt": date},
            )

            # Prodsys: filter by GSTRP (tanggal mulai produksi)
            prd_df = pd.read_sql_query(
                """
                SELECT AUFNR, MATNR,
                       SUM(CONVM)  AS qty_prd,
                       MAX(MEINS)  AS meins,
                       MAX(WERKS)  AS werks,
                       MAX(AUART)  AS auart,
                       SUM(ZDATA1) AS cnt
                FROM prodsys
                WHERE GSTRP = :dt
                GROUP BY AUFNR, MATNR
                """,
                conn, params={"dt": date},
            )

            if sap_df.empty and prd_df.empty:
                return jsonify({"data": [], "summary": {
                    "total": 0, "match": 0, "sap_lebih": 0,
                    "prodsys_lebih": 0, "hanya_sap": 0, "hanya_prodsys": 0,
                }})

            # Full outer join via pandas
            merged = pd.merge(sap_df, prd_df, on=["AUFNR", "MATNR"], how="outer", suffixes=("_s", "_p"))
            merged["qty_sap"] = merged["qty_sap"].fillna(0).round(3)
            merged["qty_prd"] = merged["qty_prd"].fillna(0).round(3)
            merged["cnt"]     = merged["cnt"].fillna(0).astype(int)
            # AUART: prefer CAUFV (via SAP side), fallback to Prodsys side
            auart_s = merged.get("auart_s", merged.get("auart", ""))
            auart_p = merged.get("auart_p", pd.Series([""] * len(merged)))
            merged["auart"] = auart_s.replace("", None).fillna(auart_p).fillna("")
            merged["werks"] = merged["werks_s"].fillna(merged["werks_p"]).fillna("")
            merged["meins"] = merged["meins_s"].fillna(merged["meins_p"]).fillna("")
            drop_cols = [c for c in ["auart_s","auart_p","werks_s","werks_p","meins_s","meins_p"] if c in merged.columns]
            if drop_cols:
                merged.drop(columns=drop_cols, inplace=True)

            # Lookup material descriptions
            matnrs = tuple(merged["MATNR"].dropna().unique().tolist())
            if matnrs:
                ph = ",".join(["?"] * len(matnrs))
                makt_rows = conn.execute(f"SELECT MATNR, MAKTX FROM makt WHERE MATNR IN ({ph})", matnrs).fetchall()
                makt_map = {r["MATNR"]: r["MAKTX"] for r in makt_rows}
                merged["maktx"] = merged["MATNR"].map(makt_map).fillna("")
            else:
                merged["maktx"] = ""

            # Selisih & status (SAP positive, Prodsys positive → selisih = SAP - Prodsys)
            merged["selisih"] = (merged["qty_sap"] - merged["qty_prd"]).round(3)

            def status(row):
                s, has_s, has_p = row["selisih"], row["qty_sap"] != 0, row["qty_prd"] != 0
                if not has_s and has_p: return "HANYA DI PRODSYS"
                if has_s and not has_p: return "HANYA DI SAP"
                if abs(s) < 0.01:      return "MATCH"
                return "SAP LEBIH BESAR" if s > 0 else "PRODSYS LEBIH BESAR"

            merged["status"] = merged.apply(status, axis=1)

            ord_ = {"SAP LEBIH BESAR":0,"PRODSYS LEBIH BESAR":1,"HANYA DI SAP":2,"HANYA DI PRODSYS":3,"MATCH":4}
            merged["_o"] = merged["status"].map(ord_)
            merged.sort_values(["_o","AUFNR","MATNR"], inplace=True)
            merged.drop("_o", axis=1, inplace=True)

            out = ["AUFNR","MATNR","maktx","auart","werks","meins","qty_sap","qty_prd","cnt","selisih","status"]
            out = [c for c in out if c in merged.columns]
            result = merged[out].fillna("")

            st = merged["status"]
            summary = {
                "total":         len(result),
                "match":         int((st=="MATCH").sum()),
                "sap_lebih":     int((st=="SAP LEBIH BESAR").sum()),
                "prodsys_lebih": int((st=="PRODSYS LEBIH BESAR").sum()),
                "hanya_sap":     int((st=="HANYA DI SAP").sum()),
                "hanya_prodsys": int((st=="HANYA DI PRODSYS").sum()),
            }

            return jsonify({"data": result.to_dict("records"), "summary": summary})

    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/export", methods=["POST"])
def export():
    try:
        payload  = request.json
        data     = payload.get("data", [])
        sel_date = payload.get("date", "comparison")

        labels = {
            "AUFNR":"Production Order","MATNR":"Material","maktx":"Description",
            "auart":"Order Type","werks":"Plant","meins":"UoM",
            "qty_sap":"Qty SAP","qty_prd":"Qty Prodsys","cnt":"Count Prodsys",
            "selisih":"Selisih","status":"Status",
        }
        df = pd.DataFrame(data).rename(columns=labels)

        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Comparison")
            from openpyxl.styles import PatternFill, Font, Alignment
            ws = writer.sheets["Comparison"]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40
                )
            cmap = {"MATCH":"C8E6C9","SAP LEBIH BESAR":"FFCDD2",
                    "PRODSYS LEBIH BESAR":"FFE0B2","HANYA DI SAP":"F8BBD0","HANYA DI PRODSYS":"FFF9C4"}
            for ri, row in enumerate(data, 2):
                color = cmap.get(row.get("status",""))
                if color:
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    for ci in range(1, len(df.columns)+1):
                        ws.cell(row=ri, column=ci).fill = fill

        out.seek(0)
        return send_file(out, as_attachment=True,
                         download_name=f"compare_{sel_date}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=8000, debug=False)
